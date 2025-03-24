import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

from openpyxl import Workbook

# -------------------------------
# CONFIGURATIONS
# -------------------------------
START_LOCATION = "YOUR_STARTING_LOCATION"  # Change this
DESTINATION = "91 Springboard, Vikhroli"
EXCEL_FILENAME = "Driving_Instructions.xlsx"
SCREENSHOT_FILENAME = "maps_screenshot.png"

# -------------------------------
# SETUP WEBDRIVER
# -------------------------------
# If chromedriver is in your PATH, you can do just webdriver.Chrome().
# Otherwise: webdriver.Chrome(executable_path="/path/to/chromedriver")
driver = webdriver.Chrome()

try:
    driver.maximize_window()

    # -------------------------------
    # 1. OPEN GOOGLE MAPS
    # -------------------------------
    driver.get("https://maps.google.com")

    # Wait until search box is present (to ensure page has loaded enough)
    WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.ID, "searchboxinput"))
    )

    # -------------------------------
    # 2. CLICK 'DIRECTIONS'
    #    XPath / ID used here
    # -------------------------------
    directions_button = driver.find_element(By.ID, "searchbox-directions")
    directions_button.click()

    # Wait a bit for the directions panel to appear
    time.sleep(2)

    # -------------------------------
    # 3. ENTER STARTING LOCATION
    #    Using an XPath that targets the first directions field
    # -------------------------------
    start_input = driver.find_element(By.XPATH, "//div[@id='directions-searchbox-0']//input")
    start_input.clear()
    start_input.send_keys(START_LOCATION)
    time.sleep(1)
    start_input.submit()  # Press Enter

    # -------------------------------
    # 4. ENTER DESTINATION
    #    Using an XPath that targets the second directions field
    # -------------------------------
    end_input = driver.find_element(By.XPATH, "//div[@id='directions-searchbox-1']//input")
    end_input.clear()
    end_input.send_keys(DESTINATION)
    time.sleep(1)
    end_input.submit()

    # -------------------------------
    # 5. SELECT THE FIRST ROUTE
    #    We wait until the route suggestions appear, then click the first route
    # -------------------------------
    WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.XPATH, "//div[contains(@class, 'section-directions-trip')]"))
    )

    first_route = driver.find_element(By.XPATH, "(//div[contains(@class,'section-directions-trip')])[1]")
    first_route.click()

    # Wait for the route details to load
    time.sleep(3)

    # -------------------------------
    # 6. EXTRACT DRIVING INSTRUCTIONS
    #    We try to find each step by its class or fallback to another class
    # -------------------------------
    # Primary XPath (older Google Maps layout)
    steps = driver.find_elements(
        By.XPATH, "//div[contains(@class, 'directions-mode-step')]//div[@class='numbered-step-content']"
    )

    # Fallback XPath (if the above yields nothing):
    if not steps:
        steps = driver.find_elements(
            By.XPATH, "//div[contains(@class,'directions-step-description')]"
        )

    instructions_list = [step.text for step in steps]

    # -------------------------------
    # 7. SAVE INSTRUCTIONS TO EXCEL
    # -------------------------------
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Driving Instructions"

    # Write header
    sheet.cell(row=1, column=1, value="Step #")
    sheet.cell(row=1, column=2, value="Instruction")

    for index, instruction in enumerate(instructions_list, start=1):
        sheet.cell(row=index + 1, column=1, value=index)
        sheet.cell(row=index + 1, column=2, value=instruction)

    workbook.save(EXCEL_FILENAME)
    print(f"Instructions saved to {EXCEL_FILENAME}")

    # -------------------------------
    # 8. TAKE A SCREENSHOT
    # -------------------------------
    driver.save_screenshot(SCREENSHOT_FILENAME)
    print(f"Screenshot saved to {SCREENSHOT_FILENAME}")

finally:
    # Close the browser
    driver.quit()